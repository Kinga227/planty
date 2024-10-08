import customtkinter as ctk
import tkinter as tk
from tkinter import font as tkfont, filedialog, StringVar, ttk
from PIL import Image
import openpyxl
from datetime import date, datetime, timedelta
from calendar import monthrange, month_name
from tkcalendar import Calendar


def load_data():
    path = ".\\plants.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    
    list_values = list(sheet.values)
    headers = list_values[0]
    plant_data = list_values[1:]

    return headers, plant_data


def display_plant_data(headers, plant_data):
    new_plant_button.pack(expand=True, fill='y', pady=(0, 5))

    for widget in content_frame.winfo_children():
        widget.destroy()

    def calculate_next_watering_date(plant):
        last_watered_date = plant[3]
        if isinstance(last_watered_date, str):
            last_watered_date = datetime.strptime(last_watered_date, "%Y-%m-%d %H:%M:%S").date()

        watering_interval = plant[2]
        if "naponta" in watering_interval:
            interval_days = int(watering_interval.split()[0])
        elif "hetente" in watering_interval:
            interval_days = int(watering_interval.split()[0]) * 7
        else:
            interval_days = 0  # default value if no interval is specified
        
        if last_watered_date:
            next_watering_date = last_watered_date + timedelta(days=interval_days)
            return next_watering_date
        else:
            return datetime.now().date()
        
    sorted_plants = sorted(plant_data, key=calculate_next_watering_date)

    row_number = 0
    for index, plant in enumerate(sorted_plants):
        if (plant[5] == 0 and active_button == mine_button) or (plant[5] == 1 and active_button == others_button):
            border_frame = ctk.CTkFrame(content_frame, fg_color="#D3D3D3", width=220, height=205)
            border_frame.grid(row=row_number, column=0, sticky="ew", padx=(75, 95), pady=(5, 0))
            
            plant_frame = ctk.CTkFrame(content_frame, fg_color="#FFFFFF")
            plant_frame.grid(row=row_number, column=0, sticky="ew", padx=(70, 100), pady=5)

            # grid configuration for the plant frame
            plant_frame.grid_columnconfigure(0, weight=0)
            plant_frame.grid_columnconfigure(1, weight=1)
            plant_frame.grid_columnconfigure(2, weight=0)

            # image of plant
            plant_img = ctk.CTkImage(light_image=Image.open(f"images\\{plant[0]}"),
                                    dark_image=Image.open(f"images\\{plant[0]}"),
                                    size=(150, 150))

            plant_label_img = ctk.CTkLabel(plant_frame, image=plant_img, text="")
            plant_label_img.grid(row=0, column=0, rowspan=5, padx=50, pady=25, sticky="n")

            # watering button
            watering_icon = ctk.CTkImage(light_image=Image.open("images\\light_mode_watering_can.png"),
                                         dark_image=Image.open("images\\light_mode_watering_can.png"),
                                         size=(55, 55))
            watering_button = ctk.CTkButton(plant_frame, text="", image=watering_icon, width=55, height=55,
                                            command=lambda p=plant, i=index: open_calendar_popup(p, i),
                                            fg_color="#FFFFFF",
                                            border_width=0)
            watering_button.grid(row=0, column=2, padx=(10, 25), pady=(10, 0), sticky="ne")

            # handle "Ontozve volt" field
            last_watered_date = plant[3]
            if isinstance(last_watered_date, str):
                last_watered_date = datetime.strptime(last_watered_date, "%Y-%m-%d %H:%M:%S").date()
            
            watering_interval = plant[2]
            if "nap" in watering_interval:
                interval_days = int(watering_interval.split()[0])
            elif "het" in watering_interval:
                interval_days = int(watering_interval.split()[0]) * 7
            else:
                interval_days = 0  # default value if no interval is specified
            
            today_date = datetime.now().date()
            
            if last_watered_date:
                # calculate next watering date
                next_watering_date = last_watered_date + timedelta(days=interval_days)
                
                # ensure next_watering_date is a date object
                if isinstance(next_watering_date, datetime):
                    next_watering_date = next_watering_date.date()
                    
                if isinstance(last_watered_date, datetime):
                    last_watered_date = last_watered_date.date()
                
                # calculate days until next watering
                days_until_next_watering = (next_watering_date - today_date).days

                if days_until_next_watering == 0:
                    next_watering_text = "ma"
                    next_watering_color = "#670505"
                elif days_until_next_watering < 0:
                    days_ago = -days_until_next_watering
                    weeks_ago = days_ago // 7
                    days_ago %= 7
                    if weeks_ago > 0 and days_ago > 0:
                        next_watering_text = f"{weeks_ago} héttel és {days_ago} nappal ezelőtt"
                    elif weeks_ago > 0:
                        next_watering_text = f"{weeks_ago} héttel ezelőtt"
                    else:
                        next_watering_text = f"{days_ago} nappal ezelőtt"
                    next_watering_color = "#FF0000" # red for overdue watering
                else:
                    if days_until_next_watering < 7:
                        next_watering_text = f"{days_until_next_watering} nap múlva"
                    else:
                        weeks = days_until_next_watering // 7
                        days = days_until_next_watering % 7
                        if days > 0:
                            next_watering_text = f"{weeks} hét és {days} nap múlva"
                        else:
                            next_watering_text = f"{weeks} hét múlva"
                    next_watering_color = "#000000" # black for upcoming watering

                # calculate time since last watering
                days_since_last_watering = (today_date - last_watered_date).days
                if days_since_last_watering == 0:
                    last_watered_text = "ma"
                elif days_since_last_watering < 7:
                    last_watered_text = f"{days_since_last_watering} napja"
                else:
                    weeks_since_last_watering = days_since_last_watering // 7
                    days = days_since_last_watering % 7
                    if days > 0:
                        last_watered_text = f"{weeks_since_last_watering} hete és {days} napja"
                    else:
                        last_watered_text = f"{weeks_since_last_watering} hete"

            else:
                next_watering_text = "ma"
                next_watering_color = "#670505" # dark red for today
                last_watered_text = "n/a"

            details = [
                f"{plant[1]}",
                f"{headers[3]}: {last_watered_text}",
                f"Következő öntözés: {next_watering_text}",
                f"{headers[4]}: {plant[4]}",
            ]

            # defining font styles for details
            font_bigger_bold = tkfont.Font(family="Istok Web", size=22, weight="bold")
            font_bold = tkfont.Font(family="Istok Web", size=18, weight="bold")
            font_regular = tkfont.Font(family="Istok Web", size=18, weight="normal")

            # applying font styles
            label_styles = [
                font_bigger_bold,   # name
                font_regular,       # watered
                font_bold,          # next watering
                font_regular        # notes
            ]
            label_colors = [
                "#0A4714",          # name color
                "#000000",          # last watered color
                next_watering_color,  # next watering color
                "#000000"           # notes color
            ]

            for i, (detail, style, color) in enumerate(zip(details, label_styles, label_colors)):
                pady_value = 0
                if i == 0:  # above name
                    pady_value = (20, 0)
                elif i == len(details) - 1:  # below notes
                    pady_value = (0, 10)
                else:   # between details
                    pady_value = (0, 0)

                tk.Label(plant_frame, text=detail, font=style, bg="#FFFFFF", fg=color).grid(
                            row=i, column=1, sticky="w", padx=5, pady=pady_value)

            row_number += 1


def open_add_new_plant_form():
    for widget in content_frame.winfo_children():
        widget.destroy()

    global upload_image_button, name_entry, watering_spinbox, watering_var, notes_textbox, add_button

    # remove 'Új hozzáadása' button
    new_plant_button.pack_forget()

    # shadow for form
    form_frame = ctk.CTkFrame(content_frame, fg_color="#D3D3D3", width=220, height=205)
    form_frame.grid(row=0, column=0, padx=(75, 95), pady=(70, 0), sticky="ew")

    form_frame.grid_rowconfigure(0, weight=1)
    form_frame.grid_columnconfigure(0, weight=1)

    # frame for the form
    inner_frame = ctk.CTkFrame(form_frame, fg_color="#FFFFFF", width=220, height=205)
    inner_frame.grid(row=0, column=0, padx=(0, 5), pady=(0, 5), sticky="ew")

    inner_frame.grid_rowconfigure(0, weight=0)
    inner_frame.grid_rowconfigure(1, weight=0)
    inner_frame.grid_rowconfigure(2, weight=0)
    inner_frame.grid_rowconfigure(3, weight=0)
    inner_frame.grid_columnconfigure(0, weight=0)
    inner_frame.grid_columnconfigure(1, weight=0)
    inner_frame.grid_columnconfigure(2, weight=0)
    inner_frame.grid_columnconfigure(3, weight=0)

    # uploading image
    upload_icon = Image.open("images\\upload.png")
    ctk_icon = ctk.CTkImage(upload_icon, size=(50, 50))
    upload_image_button = ctk.CTkButton(inner_frame, text="", command=upload_image, fg_color="#E0E0E0", image=ctk_icon, width=200, height=200)
    upload_image_button.grid(row=0, column=0, padx=50, pady=70, rowspan=3, sticky="nw")

    # name of plant
    name_label = tk.Label(inner_frame, text="Név:", font=("Istok Web", 18, "normal"), bg="#FFFFFF", fg="#000000")
    name_label.grid(row=0, column=1, padx=5, pady=(85, 0), sticky="w")
    name_entry = ctk.CTkEntry(inner_frame, width=500, bg_color="#FFFFFF", fg_color="#FFFFFF", text_color="#000000")
    name_entry.grid(row=0, column=2, padx=5, pady=(70, 0), sticky="w")

    # watering
    watering_label = tk.Label(inner_frame, text="Öntözni:", font=("Istok Web", 18, "normal"), bg="#FFFFFF", fg="#000000")
    watering_label.grid(row=1, column=1, padx=5, pady=(10, 0), sticky="w")

    watering_interval_frame = ctk.CTkFrame(inner_frame, fg_color="#FFFFFF", width=500)
    watering_interval_frame.grid(row=1, column=2, padx=(5, 5), pady=(5, 0), sticky="w")

    watering_spinbox = tk.Spinbox(watering_interval_frame, from_=1, to=30, width=5, font=("Istok Web", 16))
    watering_spinbox.grid(row=0, column=0, padx=(0, 10), pady=(0, 0), sticky="w")

    watering_var = StringVar(value="hetente")
    watering_dropdown = ttk.Combobox(watering_interval_frame, textvariable=watering_var, values=["hetente", "naponta"], state="readonly", width=10, font=("Istok Web", 16))
    watering_dropdown.grid(row=0, column=1, padx=(10, 0), pady=(0, 0), sticky="w")

    # notes
    notes_label = tk.Label(inner_frame, text="Megjegyzések:", font=("Istok Web", 18, "normal"), bg="#FFFFFF", fg="#000000")
    notes_label.grid(row=2, column=1, padx=5, pady=(0, 20), sticky="w")
    notes_textbox = tk.Text(inner_frame, height=2, width=50, font=("Istok Web", 16))
    notes_textbox.grid(row=2, column=2, padx=5, pady=(0, 20), sticky="w")

    # submit button
    add_button = ctk.CTkButton(inner_frame, text="Hozzáadás", fg_color="#05240A", command=add_new_plant, font=("Istok Web", 14, "bold"),
                               height=40, width=180, corner_radius=20)
    add_button.grid(row=3, column=0, columnspan=4, pady=(10, 20), sticky="e")

    content_frame.grid_rowconfigure(0, weight=1)
    content_frame.grid_columnconfigure(0, weight=1)

    # scroll to the top of the content frame
    scroll_to_top(content_frame)


def scroll_to_top(frame):
    frame._parent_canvas.yview_moveto(0)


def upload_image():
    global uploaded_image_path
    global image_name
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*jpeg")])
    if file_path:
        uploaded_image = Image.open(file_path).resize((150, 150))
        image_name = file_path.split("/")[-1].split("\\")[-1]
        uploaded_image_path = f"images\\{image_name}"
        uploaded_image.save(uploaded_image_path, format="PNG")
        uploaded_ctk_image = ctk.CTkImage(light_image=uploaded_image, dark_image=uploaded_image, size=(150, 150))
        upload_image_button.configure(image=uploaded_ctk_image)
        upload_image_button.image = uploaded_ctk_image

    
def add_new_plant():
    plant_name = name_entry.get()
    watering = f"{watering_spinbox.get()} {watering_var.get()}"
    notes = notes_textbox.get("1.0", "end").strip()

    if not (plant_name and watering and notes and uploaded_image_path):
        return  # do nothing if any field is empty

    # save the data to the excel file
    path = ".\\plants.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    new_plant_row = [image_name, plant_name, watering, "", notes, 0 if active_button == mine_button else 1]
    sheet.append(new_plant_row)
    workbook.save(path)

    # navigate back to the plant list and refresh
    set_active_button(active_button)


class CustomCTkButton(ctk.CTkButton):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.default_fg_color = "#FFFFFF"  # default color when inactive (white)
        self.active_color = "#C7F1CE"      # color when active or hovered (light green)
        self.calendar_color = "#05240A"    # color for calendar button (dark green)

        # button appearence config
        self.configure(
            fg_color=self.default_fg_color,
            border_width=0,
            corner_radius=0,
            text_color="#000000",
            font=("Istok Web", 32, "bold"),
            command=self.on_click
        )
        # events for hover
        self.bind("<Enter>", self.on_hover)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress>", self.on_press)
        self.bind("<ButtonRelease>", self.on_release)

    def on_hover(self, event):
        self.configure(
            fg_color=self.active_color,
            border_color=self.active_color
        )

    def on_leave(self, event):
        if self == active_button:
            self.configure(fg_color=self.active_color)  # keep color for active button
        elif self == calendar_button:
            self.configure(fg_color=self.calendar_color)  # keep calendar button color (dark green)
        else:
            self.configure(fg_color=self.default_fg_color)  # keep default button color (white)

    def on_release(self, event):
        if self == active_button:
            self.configure(fg_color=self.active_color)  # keep color for active button
        elif self == calendar_button:
            self.configure(fg_color=self.calendar_color)  # keep calendar button color (dark green)
        else:
            self.configure(fg_color=self.default_fg_color)  # keep default button color (white)

    def on_click(self):
        # set as active
        self.configure(fg_color=self.active_color)
        set_active_button(self)

    def on_press(self, event):
        # set to active color
        self.configure(fg_color=self.active_color)

    def set_active(self):
        # set to active color
        self.configure(fg_color=self.active_color)

    def set_inactive(self):
        if self == calendar_button:
            self.configure(fg_color=self.calendar_color)  # set color for calendar button
        else:
            self.configure(fg_color=self.default_fg_color)  # set color for other buttons


def set_active_button(button):
    global active_button
    for btn in [mine_button, others_button, calendar_button]:
        if btn == button:
            btn.set_active()
        else:
            btn.set_inactive()
    active_button = button  # update active button
    
    # clear content frame
    for widget in content_frame.winfo_children():
        widget.destroy()

    # update the content based on active button
    if active_button in [mine_button, others_button]:
        headers, plant_data = load_data()
        display_plant_data(headers, plant_data)

        if active_button == mine_button:
            new_plant_button.pack(expand=True, fill='y', pady=(0, 5))
    else:
        update_content(active_button)


def display_calendar():
    new_plant_button.pack_forget()

    # clear content frame
    content_frame.update_idletasks()
    for widget in content_frame.winfo_children():
        widget.destroy()

    header_frame = ctk.CTkFrame(content_frame, fg_color="#C7F1CE")
    header_frame.grid(row=0, column=0, pady=10, sticky="ew")

    header_frame.columnconfigure(0, weight=1)
    header_frame.columnconfigure(1, weight=1)
    header_frame.columnconfigure(2, weight=1)

    prev_button = ctk.CTkButton(header_frame, text="<", command=lambda: change_month(-1),
                                width=50, height=50, font=("Istok Web", 22), fg_color="#05240A")
    prev_button.grid(row=0, column=0, sticky="e")

    month_label = ctk.CTkLabel(header_frame, text=f"{month_name[selected_month]} {selected_year}",
                               font=("Istok Web", 26, "bold"), text_color="#000000")
    month_label.grid(row=0, column=1, pady=(20, 10), sticky="ew")

    next_button = ctk.CTkButton(header_frame, text=">", command=lambda: change_month(1),
                                width=50, height=50, font=("Istok Web", 22), fg_color="#05240A")
    next_button.grid(row=0, column=2, sticky="w")

    header_frame.columnconfigure(1, weight=1)

    # frame of calendar
    days_frame = ctk.CTkFrame(content_frame, fg_color="#FFFFFF")
    days_frame.grid(row=1, column=0, padx=260, pady=(30, 0), sticky="nsew")

    content_frame.grid_rowconfigure(1, weight=1)
    content_frame.grid_columnconfigure(0, weight=1)

    days_frame.grid_rowconfigure(0, weight=0)
    days_frame.grid_rowconfigure(1, weight=1)

    cell_size = 80
    
    for col in range(7):
        days_frame.grid_columnconfigure(col, minsize=cell_size, weight=1, uniform="day")
    for row in range(6):
        days_frame.grid_rowconfigure(row + 1, minsize=cell_size, weight=1)

    # name of days
    for i, day_name in enumerate(["Hétfő", "Kedd", "Szerda", "Csütörtök", "Péntek", "Szombat", "Vasárnap"]):
        day_label = ctk.CTkLabel(days_frame, text=day_name, font=("Istok Web", 18), text_color="#000000")
        day_label.grid(row=0, column=i, padx=10, pady=10, sticky="nsew")

    first_day_of_month, days_in_month = monthrange(selected_year, selected_month)
    row = 1
    col = first_day_of_month
    today = datetime.now()

    for day in range(1, days_in_month + 1):
        plant_count = count_plants_on_day(selected_year, selected_month, day)
        
        # get cell color
        if selected_year == today.year and selected_month == today.month and day == today.day:
            cell_bg_color = "#DDDDDD"  # today
            text_color = "#000000"
        elif plant_count > 0:
            cell_bg_color = "#C22124" if (today.year == selected_year and today.month == selected_month and day < today.day) else "#2E953F"
            text_color = "#FFFFFF"
        else:
            cell_bg_color = "#FFFFFF"  # default
            text_color = "#A4A4A4"
        
        # create day cell
        day_frame = ctk.CTkFrame(days_frame, fg_color=cell_bg_color, height=cell_size, width=cell_size)
        day_frame.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
        
        inner_frame = ctk.CTkFrame(day_frame, fg_color=cell_bg_color)
        inner_frame.pack(expand=True, fill="both")

        # day number label
        day_number_label = ctk.CTkLabel(inner_frame, text=str(day), font=("Istok Web", 14, "bold"), text_color=text_color)
        day_number_label.pack(anchor="nw", padx=5, pady=5)

        # plant count label
        if plant_count > 0:
            day_count_label = ctk.CTkLabel(inner_frame, text=f"{plant_count}", font=("Istok Web", 22, "bold"), text_color=text_color)
            day_count_label.pack(expand=True, fill="both", pady=(0, 10))

        col += 1
        if col > 6:
            col = 0
            row += 1

    content_frame.update_idletasks()


def change_month(direction):
    global selected_month, selected_year
    selected_month += direction
    if selected_month < 1:
        selected_month = 12
        selected_year -= 1
    elif selected_month > 12:
        selected_month = 1
        selected_year += 1

    display_calendar()

def count_plants_on_day(year, month, day):
    _, plant_data = load_data()
    count = 0
    target_date = date(year, month, day)
    # get watering dates
    for plant in plant_data:
        if plant[3]:
            last_watered_date = datetime.strptime(str(plant[3]), "%Y-%m-%d %H:%M:%S").date()
            interval_days = int(plant[2].split()[0])

            if "het" in plant[2]:
                interval_days *= 7

            next_watering_date = last_watered_date + timedelta(days=interval_days)
            if next_watering_date == target_date:
                count += 1
    return count


def update_content(button):
    global active_button
    active_button = button
    if button == calendar_button:
        display_calendar()
    else:
        display_plant_data(headers, plant_data)


def update_watering_date(plant_index, new_date):
    global plant_data
    plant_data[plant_index] = list(plant_data[plant_index])

    # update date when it was watered
    plant_data[plant_index][3] = new_date

    # save data to excel
    path = ".\\plants.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    
    # find index of plant
    row_index = plant_index + 2
    
    # update the cell with the new date
    sheet.cell(row=row_index, column=4, value=new_date.strftime("%Y-%m-%d %H:%M:%S"))

    workbook.save(path)

    # refresh the displayed data
    display_plant_data(headers, plant_data)


def open_calendar_popup(plant, plant_index):
    popup = tk.Toplevel(root)
    popup.title("Válaszd ki az utolsó öntözés dátumát")

    popup_width = 700
    popup_height = 500

    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()

    x = screen_width // 2
    y = (screen_height // 2) - (popup_height // 2)

    popup.geometry(f"{popup_width}x{popup_height}+{x}+{y}")

    popup.grid_rowconfigure(0, weight=1)
    popup.grid_rowconfigure(1, weight=0)
    popup.grid_columnconfigure(0, weight=1)

    cal = Calendar(popup,
                   selectmode="day",
                   year=datetime.now().year,
                   month=datetime.now().month,
                   day=datetime.now().day,
                   background='#C7F1CE',
                   foreground='#000000',
                   selectbackground='#05240A',
                   selectforeground='#FFFFFF',
                   bordercolor='#C7F1CE',
                   headersbackground='#C7F1CE',
                  )
    cal.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    cal.config(cursor="hand2")

    def on_date_selected():
        selected_date = cal.selection_get()
        update_watering_date(plant_index, selected_date)
        popup.destroy()

    select_button = ctk.CTkButton(popup, text="Kiválaszt", command=on_date_selected,
                                  fg_color="#05240A", height=35, width=120, corner_radius=20, font=(("Istok Web", 12, "bold")))
    select_button.grid(row=1, column=0, pady=10, sticky="s")

    popup.update_idletasks()


root = ctk.CTk()
root.title("Planty")

# retrieve screen size
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.geometry(f"{screen_width}x{screen_height-80}+-8+0")

# main frame
main_frame = ctk.CTkFrame(root, fg_color="#C7F1CE")
main_frame.pack(fill="both", expand=True)

# frame for navigation buttons
nav_frame = ctk.CTkFrame(main_frame, fg_color="#C7F1CE", height=50)
nav_frame.grid(row=0, column=0, sticky="ew")

# calculate button sizes based on screen size
button_height_ratio = 114 / 1024
button_height = int(screen_height * button_height_ratio)

# set button widths
button_width_ratio = 640 / 1440
calendar_width_ratio = 160 / 1440

buttons_width = int(screen_width * button_width_ratio)
calendar_button_width = button_height  # make calendar button square

# navigation buttons
mine_button = CustomCTkButton(main_frame, text="Saját", command=lambda: update_content(mine_button))
mine_button.grid(row=0, column=0, sticky="nsew")

others_button = CustomCTkButton(main_frame, text="Más", command=lambda: update_content(others_button))
others_button.grid(row=0, column=1, sticky="nsew")

calendar_icon = ctk.CTkImage(light_image=Image.open("images\\black_calendar.png"),
                                dark_image=Image.open("images\\white_calendar.png"),
                                size=(40, 40))

calendar_button = CustomCTkButton(main_frame, text="", image=calendar_icon, fg_color="#05240A", command=lambda: update_content(calendar_button))
calendar_button.grid(row=0, column=2, sticky="nsew")

# set sizes
mine_button.configure(height=button_height, width=buttons_width)
others_button.configure(height=button_height, width=buttons_width)
calendar_button.configure(height=calendar_button_width, width=calendar_button_width)  # ensure square dimensions

# content frame
content_frame = ctk.CTkScrollableFrame(main_frame, fg_color="#C7F1CE")
content_frame.grid(row=1, column=0, sticky="nsew", columnspan=3)  # all 3 columns
content_frame.grid_rowconfigure(0, weight=1)
content_frame.grid_columnconfigure(0, weight=1)

# frame for button to add new plant
button_frame = ctk.CTkFrame(main_frame, fg_color="#C7F1CE", height=40)
button_frame.grid(row=2, column=0, sticky="nsew", columnspan=3)

# button for adding new plant
new_plant_button = ctk.CTkButton(button_frame, text="Új hozzáadása", fg_color="#05240A",
                                height=40, width=180, corner_radius=20, font=(("Istok Web", 14, "bold")))
new_plant_button.pack(expand=True, fill='y', pady=(0, 5))
new_plant_button.configure(command=open_add_new_plant_form)

# configure sizes for main frame and content frame
main_frame.grid_rowconfigure(1, weight=1)  # allow row 1 to expand
main_frame.grid_columnconfigure(0, weight=1)  # allow column 0 to expand

# content
content_label = tk.Label(content_frame, text="", font=("Istok Web", 24), bg="#C7F1CE", fg="#000000")
content_label.pack(expand=True, fill='both')

# initial button states
active_button = mine_button  # initially mine is active
set_active_button(mine_button)  # set to active

headers, plant_data = load_data()
display_plant_data(headers, plant_data)

selected_month = datetime.now().month
selected_year = datetime.now().year

calendar_button.configure(command=display_calendar)

root.mainloop()
